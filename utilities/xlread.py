from openpyxl import load_workbook


class Utilities:
    def read_data_from_excel(file_name, sheet_name):
        global j, i
        data_list = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet_name]
        row_c = sh.max_row
        col_c = sh.max_column
        print('Maximum Rows : ', row_c, '\nColumn : ', col_c)

        for i in range(2, row_c + 1, 1):
            row = []
            for j in range(1, col_c + 1, 1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
            print(j)
            print(sh.cell(row=i, column=j).value)
        print(i)

        return data_list
